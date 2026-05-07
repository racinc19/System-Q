"""Generate CUBE_ENTRY_BOM_PRICING.xlsx — BOM for cube entry-level line (see VISUAL_CUBE_ENTRY_LEVEL.html)."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def main() -> None:
    out = Path(__file__).resolve().parent / "CUBE_ENTRY_BOM_PRICING.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "CUBE_ENTRY"

    headers = ["Qty", "Designator", "Description", "Mfr Part #", "Vendor/Notes", "Unit USD", "Ext USD"]
    rows = [
        [1, "BASE-CHS", "Main base / chassis for cube row + control strip (knob, soft fader, transport)", "Custom fab", "Wedge or flat; ties cubes mechanically", 165.00],
        [1, "CUBE-USB", "Cube 1 — 8× USB host-facing interface (hub + PCB + cube shell)", "Custom / TBD hub IC", "Computer link; same software stack as full desk", 95.00],
        [1, "CUBE-TRS", "Cube 2 — 2× 1/4″ TRS inputs (PCB + jacks + shell)", "Neutrik-style TRS", "Balanced/instrument line", 48.00],
        [1, "CUBE-XLR", "Cube 3 — 4× female XLR mic inputs (PCB + shells + shell)", "Neutrik NC3FXX or equiv", "Phantom per design", 72.00],
        [1, "CUBE-CMB", "Cube 4 — 4× combo XLR+1/4″ (PCB + shells + shell)", "Neutrik NCJ or equiv", "Mic/line/instrument", 110.00],
        [1, "MCU-BRD", "Main controller / USB audio bridge PCB (MCU, converters path)", "Custom", "Ties cubes to host + control surfaces", 135.00],
        [1, "PSU-EXT", "External PSU (5V/12V class per final design)", "Mean Well GSM/GSM-style or USB-PD brick", "Entry bundle power", 35.00],
        [1, "KNOB-6DOF", "Touch-and-turn 6DOF knob module (sensor + bearing + cap)", "Custom / 6DOF module", "SpaceMouse-class behavior", 195.00],
        [1, "FADER-LED", "Soft LED fader strip (touch + LED automation display)", "Custom", "LED follow for automation", 88.00],
        [1, "TRANS-BAR", "Transport bar — segmented touch or mechanical cluster", "Custom", "RW/FF/Stop/Play/Rec etc.", 52.00],
        [1, "WIRE-HARN", "Internal harness + USB cables + shielded audio runs", "Mixed", "Cube-to-base routing", 42.00],
        [1, "MECH-HW", "Hardware kit (screws, standoffs, feet, labels)", "Mixed", "Field service friendly", 28.00],
        [1, "DOC-PKG", "Quick-start + pinout card (print)", "—", "Ships with unit", 12.00],
    ]

    ws.append(headers)
    for row in rows:
        ws.append(row + [None])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="14532D")
        cell.alignment = Alignment(horizontal="center")

    last_data_row = 1 + len(rows)
    for r in range(2, last_data_row + 1):
        ws.cell(row=r, column=7, value=f"=A{r}*F{r}")
        ws.cell(row=r, column=6).number_format = "$#,##0.00"
        ws.cell(row=r, column=7).number_format = "$#,##0.00"

    subtotal_row = last_data_row + 1
    ws.cell(row=subtotal_row, column=5, value="Estimated subtotal (one CUBE ENTRY bundle)").font = Font(bold=True)
    ws.cell(row=subtotal_row, column=7, value=f"=SUM(G2:G{last_data_row})").font = Font(bold=True)
    ws.cell(row=subtotal_row, column=7).number_format = "$#,##0.00"

    for col, width in {"A": 8, "B": 14, "C": 56, "D": 28, "E": 46, "F": 12, "G": 12}.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
