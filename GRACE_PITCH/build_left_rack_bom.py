from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def main() -> None:
    out = Path(__file__).resolve().parent / "LEFT_RACK_BOM_PRICING.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "LEFT_BOM"

    headers = ["Qty", "Designator", "Description", "Mfr Part #", "Vendor/Notes", "Unit USD", "Ext USD"]
    rows = [
        [1, "CH-L", "Left rack enclosure with slide-out rear panel + rails + side cable channels", "Custom fab", "Per cad/system_q_left_rack_slideout.step", 450.00],
        [1, "SRV-PNL", "Perforated slide-out rear service panel assembly", "Custom fab", "Panel + rails + captive hardware set", 140.00],
        [1, "PSU-24", "AC/DC 24V front-end power supply, enclosed", "Mean Well LRS-150-24", "Main front-end feed for split rail architecture", 32.00],
        [2, "DC15-A/B", "Isolated DC/DC dual output +15V/-15V modules", "TRACO TEN 40-2415N", "Analog rail generation (headroom)", 58.00],
        [1, "PH-48", "48V phantom power module", "FiveFish Phantom Gen", "Mic phantom rail source", 29.00],
        [12, "V1-V12", "12AX7 / ECC83 dual triode tubes", "JJ ECC83S (or equivalent)", "Shared tube bank for preamp/comp/EQ tube path", 22.00],
        [12, "SK1-SK12", "Noval 9-pin tube sockets, chassis or PCB mount", "Belton 9-pin (or equivalent)", "Match mount style to final tube board", 6.50],
        [1, "HV-B+", "Tube B+ supply assembly (HV transformer/SMPS + rect/filter)", "TBD by tube circuit", "Target 200-300V class, budget placeholder", 180.00],
        [1, "HTR-12V", "Tube heater supply, regulated", "TBD by tube circuit", "12.6V class heater rail for 12 tubes", 65.00],
        [2, "XLR5-M/F", "5-pin XLR transfer connectors (male + female panel)", "Neutrik-style", "2x balanced analog channels transfer", 14.00],
        [1, "RJ45-1", "RJ45 panel connector", "Neutrik EtherCON-style", "System/network/service link", 9.00],
        [1, "CARD-SET", "6-card electronics set (5 processing + converter)", "Custom", "Prototype estimate for assembled PCB set", 1500.00],
        [1, "IO-CONN", "Other connector set (TRS/USB/power entry)", "Mixed", "Remaining back/side panel connector hardware", 220.00],
        [1, "MECH-HW", "Mechanical hardware set (standoffs, screws, rails, handle, tie points)", "Mixed", "Stainless hardware + cable management", 85.00],
        [1, "FAN-SET", "Cooling set (fans, guards, filters)", "Mixed", "Thermal control for tube + PSU zones", 95.00],
    ]

    ws.append(headers)
    for row in rows:
        ws.append(row + [None])

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center")

    last_data_row = 1 + len(rows)
    for r in range(2, last_data_row + 1):
        ws.cell(row=r, column=7, value=f"=A{r}*F{r}")
        ws.cell(row=r, column=6).number_format = "$#,##0.00"
        ws.cell(row=r, column=7).number_format = "$#,##0.00"

    subtotal_row = last_data_row + 1
    ws.cell(row=subtotal_row, column=5, value="Estimated subtotal (one LEFT rack)").font = Font(bold=True)
    ws.cell(row=subtotal_row, column=7, value=f"=SUM(G2:G{last_data_row})").font = Font(bold=True)
    ws.cell(row=subtotal_row, column=7).number_format = "$#,##0.00"

    for col, width in {"A": 8, "B": 14, "C": 56, "D": 28, "E": 46, "F": 12, "G": 12}.items():
        ws.column_dimensions[col].width = width

    ws.freeze_panes = "A2"
    wb.save(out)
    print(out)


if __name__ == "__main__":
    main()
