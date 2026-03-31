from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from pathlib import Path

out = Path(__file__).resolve().parent / "RACKS_BOM_PRICING_COMBINED.xlsx"
wb = Workbook()

left_rows = [
    [1, "CH-L", "Left rack enclosure with slide-out rear panel + rails + side cable channels", "Custom fab", "Per cad/system_q_left_rack_slideout.step", 450.00],
    [1, "PSU-24", "AC/DC 24V front-end power supply, enclosed", "Mean Well LRS-150-24", "Main front-end feed for split rail architecture", 32.00],
    [2, "DC15-A/B", "Isolated DC/DC dual output +15V/-15V modules", "TRACO TEN 40-2415N", "Analog rail generation (headroom)", 58.00],
    [1, "PH-48", "48V phantom power module", "FiveFish Phantom Gen", "Mic phantom rail source", 29.00],
    [12, "V1-V12", "12AX7 / ECC83 dual triode tubes", "JJ ECC83S (or equivalent)", "Shared tube bank for preamp/comp/EQ tube path", 22.00],
    [12, "SK1-SK12", "Noval 9-pin tube sockets, chassis or PCB mount", "Belton 9-pin (or equivalent)", "Match mount style to final tube board", 6.50],
    [1, "HV-B+", "Tube B+ supply assembly (HV transformer/SMPS + rect/filter)", "TBD by tube circuit", "Target 200-300V class, budget placeholder only", 180.00],
    [1, "HTR-12V", "Tube heater supply, regulated", "TBD by tube circuit", "12.6V class heater rail for 12 tubes", 65.00],
    [1, "CARD-SET", "6-card electronics set (5 processing + converter)", "Custom", "Prototype estimate for assembled PCB set", 1500.00],
    [1, "IO-CONN", "I/O connector set (XLR/TRS/USB/power entry)", "Mixed", "Back/side panel connector hardware", 220.00],
    [1, "MECH-HW", "Mechanical hardware set (standoffs, screws, rails, handle, tie points)", "Mixed", "Stainless hardware + cable management", 85.00],
    [1, "FAN-SET", "Cooling set (fans, guards, filters)", "Mixed", "Thermal control for tube + PSU zones", 95.00],
]

right_rows = [
    [1, "CH-R", "Right rack enclosure / frame assembly", "Custom fab", "Right-zone mechanical with internal assembly mounts", 430.00],
    [1, "MB-ASSY", "Motherboard + CPU + RAM + NVMe", "Build-specific", "Main compute assembly", 900.00],
    [1, "GPU-1", "Graphics card (UI/meters/display)", "Build-specific", "Sized to thermal and UI requirements", 450.00],
    [1, "PSU-R", "Compute PSU (ATX/SFX class)", "Build-specific", "For motherboard + GPU + digital sections", 140.00],
    [1, "DIG-RXTX", "Digital I/O board set (USB bridge, 32->12 return, optional 12 out)", "Custom", "Per CARD_LAYOUT right-zone narrative", 480.00],
    [1, "MON-BUS", "Stereo bus + monitor matrix PCB set", "Custom", "Source select, blend logic, monitor path control", 420.00],
    [3, "HP-AMP1..3", "Headphone amplifier modules", "Build-specific", "Three discrete headphone feeds", 85.00],
    [1, "SPK-SW", "Speaker select / relay board", "Build-specific", "Three monitor/speaker destinations", 120.00],
    [1, "IO-CONN-R", "I/O connector set (balanced in/out, USB/system, service)", "Mixed", "Back/side panel connector hardware", 240.00],
    [1, "MECH-HW-R", "Mechanical hardware set (standoffs, screws, shields, cable mgmt)", "Mixed", "Chassis hardware + cable retention", 95.00],
    [1, "FAN-SET-R", "Cooling set (fans, guards, filters)", "Mixed", "Thermal control for compute + analog sections", 110.00],
]


def make_sheet(name, rows, accent):
    ws = wb.create_sheet(name)
    headers = ["Qty", "Designator", "Description", "Mfr Part #", "Vendor/Notes", "Unit USD", "Ext USD"]
    ws.append(headers)
    for r in rows:
        ws.append(r + [None])
    for c in ws[1]:
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=accent)
        c.alignment = Alignment(horizontal="center")
    last = 1 + len(rows)
    for row in range(2, last + 1):
        ws.cell(row=row, column=7, value=f"=A{row}*F{row}")
        ws.cell(row=row, column=6).number_format = "$#,##0.00"
        ws.cell(row=row, column=7).number_format = "$#,##0.00"
    sub = last + 1
    ws.cell(row=sub, column=5, value="Estimated subtotal").font = Font(bold=True)
    ws.cell(row=sub, column=7, value=f"=SUM(G2:G{last})").font = Font(bold=True)
    ws.cell(row=sub, column=7).number_format = "$#,##0.00"
    for col, width in {"A": 8, "B": 14, "C": 52, "D": 28, "E": 44, "F": 12, "G": 12}.items():
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"
    return ws, sub


wb.remove(wb.active)
left_ws, left_sub = make_sheet("LEFT_BOM", left_rows, "1F4E78")
right_ws, right_sub = make_sheet("RIGHT_BOM", right_rows, "7A1F4E")

sum_ws = wb.create_sheet("SUMMARY")
sum_ws.append(["Section", "Amount USD"])
sum_ws["A1"].font = Font(bold=True, color="FFFFFF")
sum_ws["B1"].font = Font(bold=True, color="FFFFFF")
sum_ws["A1"].fill = PatternFill("solid", fgColor="1F2937")
sum_ws["B1"].fill = PatternFill("solid", fgColor="1F2937")
sum_ws.append(["LEFT subtotal", f"=LEFT_BOM!G{left_sub}"])
sum_ws.append(["RIGHT subtotal", f"=RIGHT_BOM!G{right_sub}"])
sum_ws.append(["Combined total", "=B2+B3"])
for r in (2, 3, 4):
    sum_ws.cell(row=r, column=2).number_format = "$#,##0.00"
sum_ws["A4"].font = Font(bold=True)
sum_ws["B4"].font = Font(bold=True)
sum_ws.column_dimensions["A"].width = 22
sum_ws.column_dimensions["B"].width = 14

wb.save(out)
print(out)
