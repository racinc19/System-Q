"""Write Editor button reference to Desktop as .xlsx and .csv for Google Sheets import."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


def main() -> tuple[Path, Path]:
    desk = Path(r"C:\Users\racin\Desktop")
    base = "DESKTOP_GOOGLE_SHEETS__System_Q_Editor_All_Buttons"
    out_xlsx = desk / f"{base}.xlsx"
    out_csv = desk / f"{base}.csv"
    rows: list[list[str]] = []

    def add(
        stage: str,
        ui_label: str,
        control_type: str,
        press: str,
        twist: str,
        dsp_or_state: str,
        notes: str = "",
    ) -> None:
        rows.append([stage, ui_label, control_type, press, twist, dsp_or_state, notes])

    for hdr, ena in [
        ("PRE", "pre_enabled — mic pre / filters insert"),
        ("HRM", "harmonics_enabled — harmonic generator insert"),
        ("GTE", "gate_enabled — gate / downward expander insert"),
        ("CMP", "comp_enabled — compressor insert"),
        ("EQ", "eq_enabled — equalizer insert"),
        ("TON", "tone_enabled — transient / exciter / saturation bundle"),
    ]:
        en_key = ena.split(" —")[0].strip()
        add(
            hdr,
            f"{hdr} header",
            "Column header",
            (
                "Cap PRESS / click when header is focused: toggles coarse insert "
                f"({en_key}) for this stage column."
            ),
            "(none)",
            ena,
            "Navigate UP from the top parameter row (or DOWN from header) to focus the header.",
        )

    add(
        "PRE",
        "TBE",
        "Cell",
        "Toggle pre tube (ch.tube). Does NOT toggle pre_enabled.",
        "(none)",
        "Tube saturation on PRE audio path.",
        "Lit styling: pre_enabled AND tube.",
    )
    add(
        "PRE",
        "LPF",
        "Cell",
        "Toggle ch.lpf_enabled.",
        "Twist: lpf_hz (log).",
        "Low-pass when enabled.",
        "",
    )
    add(
        "PRE",
        "48V",
        "Cell",
        "Toggle phantom (ch.phantom).",
        "(none)",
        "Phantom-power UI flag.",
        "",
    )
    add(
        "PRE",
        "PHS",
        "Cell",
        "Toggle polarity (ch.phase).",
        "(none)",
        "Phase invert on PRE.",
        "",
    )
    add(
        "PRE",
        "HPF",
        "Cell",
        "Toggle ch.hpf_enabled.",
        "Twist: hpf_hz (log).",
        "High-pass when enabled.",
        "",
    )

    add(
        "HRM",
        "TBE",
        "Cell",
        "Toggle harm_tube.",
        "(none)",
        "Harmonics-stage tube seasoning.",
        "Lit styling: harmonics_enabled AND tube.",
    )
    for i in range(1, 6):
        add(
            "HRM",
            f"H{i}",
            "Cell",
            f"Toggle harm_param_bypass[\"H{i}\"] (parameter bypass).",
            f"Twist: harmonics[{i - 1}] linear 0..1.",
            "Per-partial harmonic level; bypass mutes DSP contribution.",
            "",
        )

    add(
        "GTE",
        "TBE",
        "Cell",
        "Toggle gate_tube.",
        "(none)",
        "Gate path tube seasoning.",
        "Lit styling: gate_enabled AND tube.",
    )
    gt = (
        ("THR", "Twist: gate_threshold_db (dB rail)", "Bypass = ignore knob in detector."),
        ("RAT", "Twist: gate_ratio 1–20", ""),
        ("ATK", "Twist: gate_attack_ms", ""),
        ("RLS", "Twist: gate_release_ms", ""),
        ("GAN", "Twist: gate_makeup gain", ""),
        ("FRQ", "Twist: gate_center_hz — only meaningful when band path ON", "Band scopes detector to Hz±width."),
        ("WDT", "Twist: gate_width_oct — band width octave", ""),
    )
    for lab, twist_desc, note in gt:
        add(
            "GTE",
            lab,
            "Cell",
            f"Toggle gate_param_bypass[\"{lab}\"].",
            f"Twist: {twist_desc}",
            "Gate dynamics in _apply_gate; bypass skips that knob’s effect.",
            note,
        )
    add(
        "GTE",
        "BND",
        "Cell",
        "PRESS: toggle gate_band_enabled (narrow-band detector on/off). FRQ/WDT shape the keyed band.",
        "Discrete twist while BND row focused: ensures band ON + log-steps gate_center_hz (pick de-esser / band-centric frequency without leaving BND row).",
        "When ON, dynamics detector is band-limited (de-esser / frequency-selective expansion style). DSP: see _mono_for_dynamics_detector + _apply_gate.",
        "",
    )

    add(
        "CMP",
        "TBE",
        "Cell",
        "Toggle comp_tube.",
        "(none)",
        "Compressor path tube seasoning.",
        "Lit styling: comp_enabled AND tube — use CMP header to engage insert.",
    )
    for lab, twist_desc in [
        ("THR", "comp_threshold_db lin"),
        ("RAT", "comp_ratio lin"),
        ("ATK", "comp_attack_ms log"),
        ("RLS", "comp_release_ms log"),
        ("GAN", "comp_makeup lin"),
        ("FRQ", "comp_center_hz log"),
        ("WDT", "comp_width_oct lin"),
    ]:
        add(
            "CMP",
            lab,
            "Cell",
            f"Toggle comp_param_bypass[\"{lab}\"].",
            f"Twist: {twist_desc}",
            "Compressor dynamics in _apply_compressor.",
            (
                "THR bypass = compressor DSP idle (dry through path); polar idle. Grid still "
                "shows individual bypass toggles."
                if lab == "THR"
                else ""
            ),
        )
    add(
        "CMP",
        "BND",
        "Cell",
        "PRESS: toggle comp_band_enabled (frequency-scoped compressor detector / internal sidechain-ish band). FRQ/WDT set band envelope.",
        "Discrete twist on BND row: forces band ON + log-steps comp_center_hz (snap detector band tuning from the BND cell).",
        "When ON, detector input is filtered to center±width — classic multiband / sidechain-frequency behavior before ratio/GR applies.",
        "",
    )

    add(
        "EQ",
        "TBE",
        "Cell",
        "Toggle eq_tube.",
        "(none)",
        "EQ seasoning path.",
        "Lit styling: eq_enabled AND tube.",
    )
    add(
        "EQ",
        "FRQ",
        "Cell",
        "Toggle eq_param_bypass[\"FRQ\"].",
        "Twist: band frequency (multiband) or unified eq_freq.",
        "EQ spectral center.",
        "",
    )
    add(
        "EQ",
        "GAN",
        "Cell",
        "Toggle eq_param_bypass[\"GAN\"].",
        "Twist: gain_db on active band.",
        "EQ bell/shelf lift/cut.",
        "",
    )
    add(
        "EQ",
        "SHP",
        "Cell",
        "Toggle eq_param_bypass[\"SHP\"].",
        "Twist: bandwidth / width knob.",
        "Q / shelf shape depending on tier.",
        "",
    )
    add(
        "EQ",
        "BND",
        "Cell",
        (
            "PRESS: cycles multiband — off→on with ≥2 bands, add tiers up to 8, "
            "at max press turns multiband off. Adds frequency bands each step; "
            "FRQ/GAN/SHP/BD2 edits apply to the selected band tier."
        ),
        "Discrete twist while BND row focused AND multiband on: rotates eq_selected_band (which EQ band row FRQ/etc. tweak). Prime multiband on first EQ BND twist if needed.",
        "Band index N/tier UI; twisting moves edit focus between overlapping EQ bells/shelves.",
        "First multiband entry may auto-enable EQ insert.",
    )
    add(
        "EQ",
        "TRN",
        "Cell",
        'Toggle eq_param_bypass[\"TRN\"].',
        "(none)",
        "EQ-slot transient bypass flag.",
        "",
    )
    add(
        "EQ",
        "ATK",
        "Cell",
        'Toggle eq_param_bypass[\"ATK\"].',
        "(none)",
        "EQ-slot attack bypass.",
        "",
    )
    add(
        "EQ",
        "SUT",
        "Cell",
        'Toggle eq_param_bypass[\"SUT\"].',
        "(none)",
        "EQ-slot sustain bypass.",
        "",
    )
    add(
        "EQ",
        "BD2",
        "Cell",
        'Toggle eq_param_bypass[\"BD2\"].',
        "Twist: auxiliary width knob per band/stack.",
        "Extra width / curve assist.",
        "",
    )

    for lab, twist_desc in [
        ("TRN", "(none) — mostly bypass toggles tied to transient block"),
        ("XCT", "Twist: xct_amount lin"),
        ("DRV", "Twist: clr_drive lin"),
        ("FRQ", "Twist: xct_freq log"),
        ("ATK", "Twist: trn_attack lin"),
        ("SUT", "Twist: trn_sustain lin"),
        ("BND", "(none)"),
        ("BD2", "(none)"),
    ]:
        notes_extra = ""
        if lab == "BND":
            notes_extra = (
                "Pairs with transient band enable readout; PRESS flips "
                "tone_param_bypass['BND'] (per-knob bypass)."
            )
        add(
            "TON",
            lab,
            "Cell",
            f"Toggle tone_param_bypass[\"{lab}\"].",
            twist_desc,
            "Tone bundle in _apply_tone (TRN / CLR / XCT chain).",
            notes_extra,
        )

    add(
        "Editor transport",
        "SOLO",
        "Strip button",
        "Toggle solo for editor_channel (strip index).",
        "(none)",
        "Per-channel audition.",
        "Shown in editor-pane transport strip when wired.",
    )
    add(
        "Editor transport",
        "MUTE",
        "Strip button",
        "Toggle mute for editor_channel.",
        "(none)",
        "Mute channel.",
        "",
    )
    add(
        "Editor transport",
        "REC",
        "Strip button",
        "Toggle record arm for editor_channel.",
        "(none)",
        "Input record arm.",
        "",
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Editor_Controls"
    ws.append(["System Q — Editor pane unified matrix + editor transport footer"])
    ws.append(
        [
            "Source: software/system_q_console.py — _STAGE_GRID, "
            "_press_unified_editor_cell, _adjust_unified_editor_cell, "
            "_draw_editor_transport_row, _draw_focus_*"
        ]
    )
    ws.append(["Open in Google Sheets: Upload this file at https://drive.google.com → New → File upload"])
    ws.append([])
    hdr_row = [
        "Stage column",
        "Label",
        "UI type",
        "PRESS / click (same as unified cell activate)",
        "Twist (SpaceMouse)",
        "Audio / DSP or state variable",
        "Notes",
    ]
    ws.append(hdr_row)
    for r in rows:
        ws.append(r)

    widths = (14, 10, 12, 56, 40, 46, 50)
    for col_idx, wi in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = wi

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")
    for cell in ws[5]:
        cell.font = Font(bold=True)

    wb.save(out_xlsx)

    meta_lines = [
        ["System Q — Editor pane unified matrix + editor transport footer"],
        [
            "Source: software/system_q_console.py — _STAGE_GRID, "
            "_press_unified_editor_cell, _adjust_unified_editor_cell, "
            "_draw_editor_transport_row, _draw_focus_*"
        ],
        ["Import: Google Drive Upload or sheets.new → File → Import"],
        [],
    ]
    with out_csv.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f)
        for line in meta_lines:
            w.writerow(line)
        w.writerow(hdr_row)
        w.writerows(rows)

    howto = desk / "GOOGLE_SHEETS_OPEN_THIS_FIRST.txt"
    howto.write_text(
        "\n".join(
            [
                "System Q Editor — all matrix buttons",
                "",
                f"CSV (best for Sheets): {out_csv}",
                f"Excel workbook (also works): {out_xlsx}",
                "",
                "Open in Google Sheets:",
                "  1) https://drive.google.com",
                "  2) New > File upload  — pick the .csv OR .xlsx above",
                "  3) Right-click uploaded file > Open with > Google Sheets",
                "",
                "Or Sheets: sheets.new then File > Import > Upload.",
                "",
                "Your Desktop path is:",
                str(desk),
                "",
            ]
        ),
        encoding="utf-8",
    )
    print(out_csv)
    print(out_xlsx)
    print(howto)
    return out_xlsx, out_csv


if __name__ == "__main__":
    main()
