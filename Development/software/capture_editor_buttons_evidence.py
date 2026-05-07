#!/usr/bin/env python3
"""Drive the real System Q **application window** (not whole monitor), PNG per control step, xlsx.

Outputs (folder cleared each run — does not delete other Evidence packs):
  recording-environment/Evidence/editor-buttons/every_button_desktop/*.png
  recording-environment/Evidence/editor-buttons/every_button_desktop/manifest.csv
  recording-environment/Evidence/editor-buttons/every_button_desktop/Editor_Button_Verification.xlsx

Crops Tk root bbox so unified-matrix bypass/stripes remain readable vs full multi-monitor grabs.

Run from recording-environment/software:
  py -3 capture_editor_buttons_evidence.py
"""

from __future__ import annotations

import csv
import ctypes
import shutil
import sys
import time
from pathlib import Path


def _dpi_aware() -> None:
    if sys.platform != "win32":
        return
    try:
        ctypes.windll.shcore.SetProcessDpiAwareness(2)  # type: ignore[attr-defined]
    except Exception:
        try:
            ctypes.windll.user32.SetProcessDPIAware()  # type: ignore[attr-defined]
        except Exception:
            pass


def main() -> int:
    _dpi_aware()
    software = Path(__file__).resolve().parent
    repo = software.parent
    evid_parent = repo / "Evidence" / "editor-buttons"
    evid_parent.mkdir(parents=True, exist_ok=True)
    evid = evid_parent / "every_button_desktop"
    evid.mkdir(parents=True, exist_ok=True)
    # Fresh run: clear only this sweep folder (does not wipe bypass_matrix_full / other packs).
    for p in evid.iterdir():
        if p.is_file() and p.suffix.lower() in (
            ".png",
            ".xlsx",
            ".csv",
            ".txt",
        ):
            try:
                p.unlink()
            except OSError:
                pass

    sys.path.insert(0, str(software))

    import tkinter as tk
    from PIL import ImageGrab

    import system_q_console as sq

    # Mirror _adjust_unified_editor_cell: cells that accept analog twist.
    HAS_TWIST: set[tuple[str, str]] = {
        ("pre", "LPF"),
        ("pre", "HPF"),
        ("harm", "H1"),
        ("harm", "H2"),
        ("harm", "H3"),
        ("harm", "H4"),
        ("harm", "H5"),
        ("gate", "THR"),
        ("gate", "RAT"),
        ("gate", "ATK"),
        ("gate", "RLS"),
        ("gate", "GAN"),
        ("gate", "FRQ"),
        ("gate", "WDT"),
        ("comp", "THR"),
        ("comp", "RAT"),
        ("comp", "ATK"),
        ("comp", "RLS"),
        ("comp", "GAN"),
        ("comp", "FRQ"),
        ("comp", "WDT"),
        ("eq", "FRQ"),
        ("eq", "GAN"),
        ("eq", "SHP"),
        ("eq", "BD2"),
        ("tone", "DRV"),
        ("tone", "XCT"),
        ("tone", "FRQ"),
        ("tone", "ATK"),
        ("tone", "SUT"),
    }

    # Discrete-twist BND paths (not in HAS_TWIST linear set).
    HAS_DISCRETE_TWIST: set[tuple[str, str]] = {
        ("gate", "BND"),
        ("comp", "BND"),
        ("eq", "BND"),
    }

    INSERT_GETTERS = {
        "pre": lambda ch: bool(ch.pre_enabled),
        "harm": lambda ch: bool(ch.harmonics_enabled),
        "gate": lambda ch: bool(ch.gate_enabled),
        "comp": lambda ch: bool(ch.comp_enabled),
        "eq": lambda ch: bool(ch.eq_enabled),
        "tone": lambda ch: bool(ch.tone_enabled),
    }

    twist_axis = 0.55
    root = tk.Tk()
    root.geometry("1720x1040")
    root.title("System Q Console")
    app = sq.ConsoleApp(root, startup_play=False)
    root.deiconify()

    report_rows: list[
        tuple[str, str, str, str, str, str, str, str]
    ] = []  # RowID, StageUI, Label, Role, PRESS, TWIST, Screenshots, Overall

    seq = 0

    def pump() -> None:
        for _ in range(6):
            root.update_idletasks()
            root.update()

    def foreground() -> None:
        root.lift()
        try:
            root.attributes("-topmost", True)
        except tk.TclError:
            pass
        pump()
        root.focus_force()
        try:
            root.after(80, lambda: root.attributes("-topmost", False))
        except tk.TclError:
            pass
        pump()
        time.sleep(0.08)

    def desktop_grab(tag: str) -> Path:
        nonlocal seq
        foreground()
        time.sleep(0.12)
        pump()
        # Force repaint so bypass stripes / muted SHP pixels exist before crop (grab is not desktop-wide).
        app._draw_editor_controls()
        app._draw_focus()
        pump()
        root.update_idletasks()
        rx, ry = int(root.winfo_rootx()), int(root.winfo_rooty())
        rw = max(int(root.winfo_width()), 880)
        rh = max(int(root.winfo_height()), 640)
        img = ImageGrab.grab((rx, ry, rx + rw, ry + rh), all_screens=True)
        seq += 1
        out = evid / f"{seq:03d}_{tag}.png"
        img.save(out)
        return out

    def configure_editor_matrix() -> None:
        app.nav_scope = "editor"
        app.editor_nav_scope = "stage_grid"
        app.editor_channel = 0
        app.selected_channel = 0
        app._normalize_stage_selection(0)
        app.editor_unified_header_focus = False
        app._coerce_editor_nav_to_unified_stage_grid()

    def prime_session_channel() -> None:
        ch = app.engine.channels[0]
        with app.engine._lock:
            ch.pre_enabled = True
            ch.harmonics_enabled = True
            ch.gate_enabled = True
            ch.comp_enabled = True
            ch.eq_enabled = True
            ch.tone_enabled = True
            # Harmonics visible for H1..H5 twist
            for i in range(5):
                ch.harmonics[i] = 0.35
            ch.lpf_enabled = True
            ch.hpf_enabled = True
            ch.lpf_hz = 14000.0
            ch.hpf_hz = 450.0
            ch.gate_band_enabled = True
            ch.comp_band_enabled = True
            ch.gate_dyn_band_count = 2
            ch.gate_dyn_ui_band = 0
            ch.comp_dyn_band_count = 2
            ch.comp_dyn_ui_band = 0
            gb0 = ch.gate_dyn_bands[0]
            gb1 = ch.gate_dyn_bands[1]
            gb0.update(
                freq=3000.0,
                width_oct=3.9,
                threshold_db=-42.0,
                ratio=8.0,
                attack_ms=4.0,
                release_ms=130.0,
                makeup=1.0,
                enabled=True,
            )
            gb1.update(
                freq=6400.0,
                width_oct=2.2,
                threshold_db=-44.0,
                ratio=10.0,
                attack_ms=2.0,
                release_ms=180.0,
                makeup=1.0,
                enabled=True,
            )
            cb0 = ch.comp_dyn_bands[0]
            cb1 = ch.comp_dyn_bands[1]
            cb0.update(
                freq=2800.0,
                width_oct=3.5,
                threshold_db=-22.0,
                ratio=4.0,
                attack_ms=12.0,
                release_ms=110.0,
                makeup=1.0,
                enabled=True,
            )
            cb1.update(
                freq=9000.0,
                width_oct=2.0,
                threshold_db=-26.0,
                ratio=6.0,
                attack_ms=6.0,
                release_ms=90.0,
                makeup=1.05,
                enabled=True,
            )
            app.engine._hydrate_gate_dyn_to_scalars(ch)
            app.engine._hydrate_comp_dyn_to_scalars(ch)
            ch.eq_band_enabled = True
            ch.eq_band_count = 2
            app.eq_selected_band = 0
            b0 = app._eq_band(ch, 0)
            b1 = app._eq_band(ch, 1)
            b0.update(
                {
                    "freq": 1200.0,
                    "gain_db": 4.0,
                    "width": 1.35,
                    "type": "BELL",
                    "enabled": True,
                }
            )
            b1.update(
                {
                    "freq": 4500.0,
                    "gain_db": -2.0,
                    "width": 1.40,
                    "type": "BELL",
                    "enabled": True,
                }
            )
            app._sync_scalar_display_from_eq_band(ch)
            ch.harm_param_bypass.clear()
            ch.gate_param_bypass.clear()
            ch.comp_param_bypass.clear()
            ch.eq_param_bypass.clear()
            ch.tone_param_bypass.clear()
            ch.transient_enabled = True
            ch.saturation_enabled = True
            ch.exciter_enabled = True
            ch.trn_band_enabled = True
            ch.xct_band_enabled = True
            ch.trn_attack = 0.38
            ch.trn_sustain = 0.36
            ch.clr_drive = 0.52
            ch.xct_amount = 0.52
            ch.xct_freq = 7200.0
            ch.xct_width = 1.2
            ch.trn_freq = 2400.0
            ch.trn_width = 1.3
        app._draw_editor_controls()
        app._sync_from_engine()
        pump()

    def focus_cell(col: int, row: int) -> None:
        configure_editor_matrix()
        app.editor_stage_col = col
        app.editor_param_row = row
        app.editor_unified_header_focus = False
        sk = app._STAGE_GRID[col][0]
        app.selected_stage_key = sk
        app._unified_commit_param_row_for_col(col, row)
        app._draw_editor_controls()
        app._sync_from_engine()
        pump()

    def focus_header(col: int) -> None:
        configure_editor_matrix()
        app.editor_stage_col = col
        app.editor_unified_header_focus = True
        plist = app._STAGE_GRID[col][2]
        app.editor_param_row = min(app.editor_param_row, max(0, len(plist) - 1))
        app.selected_stage_key = app._STAGE_GRID[col][0]
        app._unified_commit_param_row_for_col(col, app.editor_param_row)
        app._draw_editor_controls()
        app._sync_from_engine()
        pump()

    def snap_pair(base: str, after_action) -> tuple[Path, Path]:
        """Full desktop before (focus) and after callback."""
        before = desktop_grab(f"{base}_before")
        after_action()
        pump()
        aft = desktop_grab(f"{base}_after")
        return before, aft

    def insert_state(sk: str):
        ch = app._current_channel()
        return INSERT_GETTERS[sk](ch)

    def stage_display(sk: str, label: str) -> str:
        ch = app._current_channel()
        s, _ = app._stage_cell_value(ch, sk, label)
        return str(s)

    def run_header(col: int, hdr: str, sk: str, row_id: str) -> None:
        nonlocal report_rows
        focus_header(col)
        pics: list[str] = []
        pics.append(desktop_grab(f"{row_id}_focus").name)
        before = insert_state(sk)

        def press_hdr() -> None:
            app._press_unified_editor_cell()

        b_path, a_path = snap_pair(row_id + "_hdr_press", press_hdr)
        pics.extend([b_path.name, a_path.name])
        after = insert_state(sk)
        press_ok = before != after
        # restore
        app._press_unified_editor_cell()
        pump()
        pics.append(desktop_grab(f"{row_id}_restored").name)
        ov = "PASS" if press_ok else "BLOCKED"
        report_rows.append((row_id, hdr, "HEADER", "matrix_header", press_ok, "n/a", ";".join(pics), ov))

    def run_matrix_cell(col: int, row: int, sk: str, label: str, row_id: str) -> None:
        nonlocal report_rows
        if sk == "tone":
            if label == "XCT":
                app.tone_editor_mode = "XCT"
            elif label == "DRV":
                app.tone_editor_mode = "CLR"
            else:
                app.tone_editor_mode = "TRN"
        focus_cell(col, row)
        pics: list[str] = []
        ch = app._current_channel()

        # ---- PRESS proof (toggle / bypass / EQ BND advance / band toggles)
        def do_press() -> None:
            app._press_unified_editor_cell()

        pre_txt = stage_display(sk, label)
        pics.append(desktop_grab(f"{row_id}_press_before").name)

        # State delta by label (programmatic, stricter than OCR on PNG)
        press_ok = False
        if label == "TBE" and sk in ("pre", "harm", "gate", "comp", "eq"):
            attr = {
                "pre": "tube",
                "harm": "harm_tube",
                "gate": "gate_tube",
                "comp": "comp_tube",
                "eq": "eq_tube",
            }[sk]
            b0 = bool(getattr(ch, attr))
            do_press()
            pump()
            press_ok = bool(getattr(ch, attr)) != b0
        elif sk == "pre" and label in ("LPF", "48V", "PHS", "HPF"):
            attr = {"LPF": "lpf_enabled", "48V": "phantom", "PHS": "phase", "HPF": "hpf_enabled"}[label]
            b0 = bool(getattr(ch, attr))
            do_press()
            pump()
            press_ok = bool(getattr(ch, attr)) != b0
        elif sk == "harm" and label.startswith("H"):
            hp = ch.harm_param_bypass
            b0 = bool(hp.get(label, False))
            do_press()
            pump()
            press_ok = bool(hp.get(label, False)) != b0
        elif sk in ("gate", "comp") and label == "BND":
            n_attr = "gate_dyn_band_count" if sk == "gate" else "comp_dyn_band_count"
            band_attr = "gate_band_enabled" if sk == "gate" else "comp_band_enabled"
            n0 = int(getattr(ch, n_attr))
            en0 = bool(getattr(ch, band_attr))
            do_press()
            pump()
            n1 = int(getattr(ch, n_attr))
            en1 = bool(getattr(ch, band_attr))
            press_ok = en0 != en1 or n0 != n1
        elif sk in ("gate", "comp") and label not in ("TBE", "BND"):
            en_attr = "gate_enabled" if sk == "gate" else "comp_enabled"
            b0 = bool(getattr(ch, en_attr))
            do_press()
            pump()
            press_ok = bool(getattr(ch, en_attr)) != b0
        elif sk == "eq":
            bp = ch.eq_param_bypass
            if label in ("FRQ", "GAN", "SHP", "TRN", "ATK", "SUT", "BD2"):
                b0 = bool(bp.get(label, False))
                do_press()
                pump()
                press_ok = bool(bp.get(label, False)) != b0
            elif label == "BND":
                n0 = int(ch.eq_band_count)
                en0 = bool(ch.eq_band_enabled)
                do_press()
                pump()
                press_ok = bool(ch.eq_band_enabled) != en0 or int(ch.eq_band_count) != n0
            elif label == "TBE":
                b0 = bool(ch.eq_tube)
                do_press()
                pump()
                press_ok = bool(ch.eq_tube) != b0
            else:
                do_press()
                pump()
                press_ok = True
        elif sk == "tone":
            tb = ch.tone_param_bypass
            b0 = bool(tb.get(label, False))
            do_press()
            pump()
            press_ok = bool(tb.get(label, False)) != b0
        else:
            do_press()
            pump()
            press_ok = pre_txt != stage_display(sk, label)

        pics.append(desktop_grab(f"{row_id}_press_after").name)

        # Undo most press actions for stable next steps
        undo_press = False
        if label == "TBE" and sk in ("pre", "harm", "gate", "comp", "eq"):
            undo_press = True
        elif sk == "pre" and label in ("LPF", "48V", "PHS", "HPF"):
            undo_press = True
        elif sk == "harm" and label.startswith("H"):
            undo_press = True
        elif sk in ("gate", "comp") and label not in ("BND",):
            undo_press = True
        elif sk == "tone":
            undo_press = True
        elif sk == "eq" and label != "BND":
            undo_press = True

        if undo_press:
            app._press_unified_editor_cell()
            pump()

        # Restore EQ tier if we incremented band count
        if sk == "eq" and label == "BND":
            with app.engine._lock:
                ch.eq_band_count = 2
                ch.eq_band_enabled = True
                app.eq_selected_band = 0
                app._sync_scalar_display_from_eq_band(ch)
            app._draw_editor_controls()
            app._sync_from_engine()
            pump()

        if sk in ("gate", "comp") and label == "BND":
            with app.engine._lock:
                if sk == "gate":
                    ch.gate_band_enabled = True
                    ch.gate_dyn_band_count = 2
                    ch.gate_dyn_ui_band = 0
                    app.engine._hydrate_gate_dyn_to_scalars(ch)
                else:
                    ch.comp_band_enabled = True
                    ch.comp_dyn_band_count = 2
                    ch.comp_dyn_ui_band = 0
                    app.engine._hydrate_comp_dyn_to_scalars(ch)
            app._draw_editor_controls()
            app._sync_from_engine()
            pump()

        # ---- TWIST proof
        focus_cell(col, row)
        twist_note = "n/a"
        twist_ok = True
        if (sk, label) in HAS_TWIST:

            def read_scalar_for_twist():
                """Return a comparable float for numeric change detection."""
                if sk == "pre" and label == "LPF":
                    return float(ch.lpf_hz)
                if sk == "pre" and label == "HPF":
                    return float(ch.hpf_hz)
                if sk == "harm" and label.startswith("H"):
                    ix = int(label[1]) - 1
                    return float(ch.harmonics[ix])
                if sk == "gate":
                    keys = {
                        "THR": ("gate_threshold_db",),
                        "RAT": ("gate_ratio",),
                        "ATK": ("gate_attack_ms",),
                        "RLS": ("gate_release_ms",),
                        "GAN": ("gate_makeup",),
                        "FRQ": ("gate_center_hz",),
                        "WDT": ("gate_width_oct",),
                    }
                    return float(getattr(ch, keys[label][0]))
                if sk == "comp":
                    keys = {
                        "THR": ("comp_threshold_db",),
                        "RAT": ("comp_ratio",),
                        "ATK": ("comp_attack_ms",),
                        "RLS": ("comp_release_ms",),
                        "GAN": ("comp_makeup",),
                        "FRQ": ("comp_center_hz",),
                        "WDT": ("comp_width_oct",),
                    }
                    return float(getattr(ch, keys[label][0]))
                if sk == "eq" and label in ("FRQ", "GAN", "SHP", "BD2"):
                    bmap = {"FRQ": "freq", "GAN": "gain_db", "SHP": "width", "BD2": "width"}
                    return float(app._eq_band(ch)[bmap[label]])
                if sk == "tone":
                    keys = {
                        "DRV": "clr_drive",
                        "XCT": "xct_amount",
                        "FRQ": "xct_freq",
                        "ATK": "trn_attack",
                        "SUT": "trn_sustain",
                    }
                    return float(getattr(ch, keys[label]))
                return 0.0

            tv0 = read_scalar_for_twist()
            pics.append(desktop_grab(f"{row_id}_twist_before").name)
            time.sleep(0.22)
            app._adjust_unified_editor_cell(twist_axis)
            pump()
            tv1 = read_scalar_for_twist()
            pics.append(desktop_grab(f"{row_id}_twist_after").name)
            twist_ok = abs(tv1 - tv0) > 1e-6
            twist_note = "twist+" if twist_ok else "no_delta"
            # restore approximately
            time.sleep(0.22)
            app._adjust_unified_editor_cell(-twist_axis * 0.92)
            pump()
        elif (sk, label) in HAS_DISCRETE_TWIST:
            if sk == "eq" and label == "BND":
                tx0 = stage_display(sk, label)
                pics.append(desktop_grab(f"{row_id}_twist_before").name)
                time.sleep(0.22)
                app._adjust_unified_editor_cell(twist_axis)
                pump()
                tx1 = stage_display(sk, label)
                pics.append(desktop_grab(f"{row_id}_twist_after").name)
                twist_ok = tx0 != tx1
                twist_note = "discrete_band" if twist_ok else "no_delta"
            elif sk in ("gate", "comp") and label == "BND":
                tx0 = stage_display(sk, label)
                pics.append(desktop_grab(f"{row_id}_twist_before").name)
                time.sleep(0.22)
                app._adjust_unified_editor_cell(twist_axis)
                pump()
                tx1 = stage_display(sk, label)
                pics.append(desktop_grab(f"{row_id}_twist_after").name)
                twist_ok = tx0 != tx1
                twist_note = "discrete_tier" if twist_ok else "no_delta"
                time.sleep(0.22)
                app._adjust_unified_editor_cell(-twist_axis * 0.9)
                pump()
            else:
                hz_attr = "gate_center_hz" if sk == "gate" else "comp_center_hz"
                hz0 = float(getattr(ch, hz_attr))
                pics.append(desktop_grab(f"{row_id}_twist_before").name)
                time.sleep(0.22)
                app._adjust_unified_editor_cell(twist_axis)
                pump()
                hz1 = float(getattr(ch, hz_attr))
                pics.append(desktop_grab(f"{row_id}_twist_after").name)
                twist_ok = abs(hz1 - hz0) > 1.0  # Hz moved meaningfully
                twist_note = "discrete_hz" if twist_ok else "no_delta"
                time.sleep(0.22)
                app._adjust_unified_editor_cell(-twist_axis * 0.9)
                pump()
        else:
            twist_note = "n/a"

        ov = "PASS" if press_ok and twist_ok else "BLOCKED"
        report_rows.append(
            (
                row_id,
                app._STAGE_GRID[col][1],
                label,
                "matrix_cell",
                "PASS" if press_ok else "BLOCKED",
                twist_note,
                ";".join(pics),
                ov,
            )
        )

    def run_transport(name: str, sel: int, row_id: str) -> None:
        nonlocal report_rows
        configure_editor_matrix()
        app.nav_scope = "editor"
        app.editor_nav_scope = "transport"
        app.editor_transport_selected = sel
        app._draw_editor_controls()
        app._sync_from_engine()
        pump()
        idx = app._active_channel_index()
        ch = app.engine.channels[idx]
        attr = {"SOLO": "solo", "MUTE": "mute", "REC": "record_armed"}[name]
        pics = [desktop_grab(f"{row_id}_focus").name]
        b0 = bool(getattr(ch, attr))

        def do_t() -> None:
            if name == "SOLO":
                app._toggle_solo(idx)
            elif name == "MUTE":
                app._toggle_mute(idx)
            else:
                app._toggle_record_arm(idx)

        p0, p1 = snap_pair(row_id + "_toggle", do_t)
        pics.extend([p0.name, p1.name])
        pump()
        b1 = bool(getattr(ch, attr))
        ok = b0 != b1
        # restore
        if ok:
            do_t()
            pump()
        pics.append(desktop_grab(f"{row_id}_restored").name)
        report_rows.append(
            (row_id, "Editor transport", name, "transport", "PASS" if ok else "BLOCKED", "n/a", ";".join(pics), "PASS" if ok else "BLOCKED")
        )

    # --- Run
    prime_session_channel()
    desktop_grab("000_suite_start_baseline")

    for col, (sk, hdr, params) in enumerate(app._STAGE_GRID):
        run_header(col, hdr, sk, f"{hdr}_HEADER")

    for col, (sk, hdr, params) in enumerate(app._STAGE_GRID):
        for row, label in enumerate(params):
            rid = f"{hdr}_{label}"
            run_matrix_cell(col, row, sk, label, rid)

    run_transport("SOLO", 0, "TRN_SOLO")
    run_transport("MUTE", 1, "TRN_MUTE")
    run_transport("REC", 2, "TRN_REC")

    desktop_grab("ZZZ_suite_end")

    root.destroy()

    # --- Workbook (green / red on Overall column)
    xlsx_path = evid / "Editor_Button_Verification.xlsx"
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill

        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        wb = Workbook()
        ws = wb.active
        ws.title = "Verification"
        ws.append(
            [
                "Row ID",
                "Stage column",
                "Label",
                "Role",
                "PRESS",
                "TWIST",
                "Screenshot files",
                "Overall",
            ]
        )
        for cell in ws[1]:
            cell.font = Font(bold=True)
        for row in report_rows:
            ws.append(list(row))
        for r in range(2, ws.max_row + 1):
            val = ws.cell(row=r, column=8).value
            ws.cell(row=r, column=8).fill = green if val == "PASS" else red
        wb.save(xlsx_path)
    except ImportError:
        csv_path = evid / "Editor_Button_Verification.csv"
        with csv_path.open("w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(["Row ID", "Stage column", "Label", "Role", "PRESS", "TWIST", "Screenshot files", "Overall"])
            w.writerows(report_rows)
        shutil.copy(csv_path, xlsx_path.with_suffix(".csv-only.txt"))
        print("openpyxl missing; wrote CSV instead:", csv_path, flush=True)

    man_path = evid / "manifest.csv"
    with man_path.open("w", newline="", encoding="utf-8-sig") as mf:
        mw = csv.writer(mf)
        mw.writerow(
            [
                "Row ID",
                "Stage column",
                "Label",
                "Role",
                "PRESS",
                "TWIST",
                "Screenshot files",
                "Overall",
            ]
        )
        mw.writerows(report_rows)

    failed = [r for r in report_rows if r[7] != "PASS"]
    print(f"Evidence dir: {evid}", flush=True)
    print(f"Workbook: {xlsx_path}", flush=True)
    print(f"manifest.csv -> {man_path}", flush=True)
    print(f"Rows PASS {len(report_rows) - len(failed)} / {len(report_rows)}", flush=True)
    if failed:
        print("BLOCKED:", flush=True)
        for r in failed:
            print(" ", r[0], r[4], r[5], flush=True)
        return 1
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
